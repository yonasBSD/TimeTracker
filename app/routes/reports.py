import csv
import io
import time
from datetime import datetime, timedelta

from flask import Blueprint, current_app, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_babel import _
from flask_login import current_user, login_required
from sqlalchemy import case, func, or_
from sqlalchemy.orm import joinedload

from app import db, log_event, track_event
from app.models import (
    Client,
    Invoice,
    Payment,
    Project,
    ProjectCost,
    ReportEmailSchedule,
    SavedReportView,
    Settings,
    Task,
    TimeEntry,
    User,
)
from app.repositories import TimeEntryRepository
from app.services.scheduled_report_service import ScheduledReportService
from app.utils.support_report_generation import record_report_generation_for_current_user
from app.utils.excel_export import create_project_report_excel, create_time_entries_excel
from app.utils.posthog_monitoring import track_error, track_export_performance, track_validation_error

# Optional PowerPoint export - only import if available
try:
    from app.utils.powerpoint_export import create_report_powerpoint

    PPTX_EXPORT_AVAILABLE = True
except ImportError:
    PPTX_EXPORT_AVAILABLE = False
    create_report_powerpoint = None

reports_bp = Blueprint("reports", __name__)
from app.utils.module_helpers import module_enabled


@reports_bp.route("/reports")
@login_required
@module_enabled("reports")
def reports():
    """Main reports page - REFACTORED to use service layer with optimized queries"""
    from app.services import ReportingService

    # Use service layer to get reports summary (optimized queries)
    reporting_service = ReportingService()
    from app.telemetry.otel_setup import business_span, record_report_generated

    with business_span("report.generate", user_id=current_user.id, report_type="summary"):
        result = reporting_service.get_reports_summary(user_id=current_user.id, is_admin=current_user.is_admin)
    record_report_generated()

    # Track report access
    log_event("report.viewed", user_id=current_user.id, report_type="summary")
    track_event(current_user.id, "report.viewed", {"report_type": "summary"})

    return render_template(
        "reports/index.html",
        summary=result["summary"],
        recent_entries=result["recent_entries"],
        comparison=result["comparison"],
    )


@reports_bp.route("/reports/week-in-review")
@login_required
@module_enabled("reports")
def week_in_review():
    """Week in review: this week's hours, top projects, billable vs non-billable."""
    from app.services import ReportingService

    reporting_service = ReportingService()
    data = reporting_service.get_week_in_review(user_id=current_user.id, is_admin=current_user.is_admin)
    if data.get("error"):
        flash(data["error"], "error")
        return redirect(url_for("reports.reports"))
    return render_template("reports/week_in_review.html", **data)


@reports_bp.route("/reports/comparison")
@login_required
@module_enabled("reports")
def comparison_view():
    """Get comparison data for reports"""
    from app.services import ReportingService

    period = request.args.get("period", "month")
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    data = ReportingService().get_comparison_data(period=period, user_id=current_user.id, can_view_all=can_view_all)
    return jsonify(data)


@reports_bp.route("/reports/project")
@login_required
@module_enabled("reports")
def project_report():
    """Project-based time report"""
    from app.services import ReportingService
    from app.utils.scope_filter import apply_project_scope_to_model

    project_id = request.args.get("project_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    user_id = request.args.get("user_id", type=int)

    projects_query = Project.query.filter_by(status="active").order_by(Project.name)
    scope_p = apply_project_scope_to_model(Project, current_user)
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    projects = projects_query.all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return render_template("reports/project_report.html", projects=projects, users=users)

    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    if user_id and not can_view_all and user_id != current_user.id:
        flash(_("You do not have permission to view other users' time entries"), "error")
        return render_template("reports/project_report.html", projects=projects, users=users)

    data = ReportingService().get_project_report_data(
        start_dt=start_dt,
        end_dt=end_dt,
        project_id=project_id,
        user_id_filter=user_id,
        current_user_id=current_user.id,
        can_view_all=can_view_all,
    )
    return render_template(
        "reports/project_report.html",
        projects=projects,
        users=users,
        entries=data["entries"],
        projects_data=data["projects_data"],
        summary=data["summary"],
        start_date=start_date,
        end_date=end_date,
        selected_project=project_id,
        selected_user=user_id,
    )


@reports_bp.route("/reports/user")
@login_required
@module_enabled("reports")
def user_report():
    """User-based time report"""
    user_id = request.args.get("user_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    project_id = request.args.get("project_id", type=int)

    # Get users for filter
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    from app.utils.scope_filter import apply_project_scope_to_model

    projects_query = Project.query.filter_by(status="active").order_by(Project.name)
    scope_p = apply_project_scope_to_model(Project, current_user)
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    projects = projects_query.all()

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return render_template("reports/user_report.html", users=users, projects=projects)

    # Get time entries
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if user_id:
        # Only allow filtering by other users if they have permission
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            # User doesn't have permission to view other users' entries
            flash(_("You do not have permission to view other users' time entries"), "error")
            return render_template("reports/user_report.html", users=users, projects=projects)

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)

    entries = (
        query.options(
            joinedload(TimeEntry.project),
            joinedload(TimeEntry.user),
        )
        .order_by(TimeEntry.start_time.desc())
        .all()
    )

    # Calculate totals
    total_hours = sum(entry.duration_hours for entry in entries)
    billable_hours = sum(entry.duration_hours for entry in entries if entry.billable)

    # Group by user
    user_totals = {}
    projects_set = set()
    users_set = set()
    for entry in entries:
        if entry.project:
            projects_set.add(entry.project.id)
        if entry.user:
            users_set.add(entry.user.id)
        username = entry.user.display_name if entry.user else "Unknown"
        if username not in user_totals:
            user_totals[username] = {
                "hours": 0,
                "billable_hours": 0,
                "entries": [],
                "user_obj": entry.user,  # Store user object for overtime calculation
            }
        user_totals[username]["hours"] += entry.duration_hours
        if entry.billable:
            user_totals[username]["billable_hours"] += entry.duration_hours
        user_totals[username]["entries"].append(entry)

    # Calculate overtime for each user
    from app.utils.overtime import calculate_period_overtime

    for username, data in user_totals.items():
        if data["user_obj"]:
            overtime_data = calculate_period_overtime(data["user_obj"], start_dt.date(), end_dt.date())
            data["regular_hours"] = overtime_data["regular_hours"]
            data["overtime_hours"] = overtime_data["overtime_hours"]
            data["undertime_hours"] = overtime_data.get("undertime_hours", 0)
            data["days_under"] = overtime_data.get("days_under", 0)
            data["days_with_overtime"] = overtime_data["days_with_overtime"]

    summary = {
        "total_hours": round(total_hours, 1),
        "billable_hours": round(billable_hours, 1),
        "users_count": len(users_set),
        "projects_count": len(projects_set),
    }

    return render_template(
        "reports/user_report.html",
        users=users,
        projects=projects,
        entries=entries,
        user_totals=user_totals,
        summary=summary,
        start_date=start_date,
        end_date=end_date,
        selected_user=user_id,
        selected_project=project_id,
    )


@reports_bp.route("/reports/export/form")
@login_required
@module_enabled("reports")
def export_form():
    """Display export form with filter options (CSV or Excel)."""
    # Get all users (for admin)
    users = []
    if current_user.is_admin:
        users = User.query.filter_by(is_active=True).order_by(User.username).all()

    # Get all active projects (scoped for subcontractors)
    from app.utils.scope_filter import apply_client_scope_to_model, apply_project_scope_to_model

    projects_query = Project.query.filter_by(status="active").order_by(Project.name)
    scope_p = apply_project_scope_to_model(Project, current_user)
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    projects = projects_query.all()

    # Get all active clients (scoped for subcontractors)
    clients_query = Client.query.filter_by(status="active").order_by(Client.name)
    scope_c = apply_client_scope_to_model(Client, current_user)
    if scope_c is not None:
        clients_query = clients_query.filter(scope_c)
    clients = clients_query.all()
    only_one_client = len(clients) == 1
    single_client = clients[0] if only_one_client else None

    # Set default date range (last 30 days)
    default_end_date = datetime.utcnow().strftime("%Y-%m-%d")
    default_start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")

    # Format from query (csv or excel) for Quick Actions consistency
    export_format = request.args.get("format", "csv").lower()
    if export_format not in ("csv", "excel"):
        export_format = "csv"

    return render_template(
        "reports/export_form.html",
        users=users,
        projects=projects,
        clients=clients,
        only_one_client=only_one_client,
        single_client=single_client,
        default_start_date=default_start_date,
        default_end_date=default_end_date,
        export_format=export_format,
    )


@reports_bp.route("/reports/export/csv")
@login_required
@module_enabled("reports")
def export_csv():
    """Export time entries as CSV with enhanced filters"""
    from app.utils.client_lock import enforce_locked_client_id

    start_time = time.time()  # Start performance tracking

    # Get all filter parameters
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)
    task_id = request.args.get("task_id", type=int)
    client_id = request.args.get("client_id", type=int)
    client_id = enforce_locked_client_id(client_id)
    billable = request.args.get("billable")  # 'yes', 'no', or 'all'
    source = request.args.get("source")  # 'manual', 'auto', or 'all'
    tags = request.args.get("tags", "").strip()

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        track_validation_error(
            current_user.id,
            "date_range",
            "Invalid date format for CSV export",
            {"start_date": start_date, "end_date": end_date},
        )
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.reports"))

    # Get time entries
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if user_id:
        # Only allow filtering by other users if they have permission
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            flash(_("You do not have permission to export other users' time entries"), "error")
            return redirect(url_for("reports.reports"))

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)

    entries = (
        query.options(
            joinedload(TimeEntry.project).joinedload(Project.client_obj),
            joinedload(TimeEntry.user),
            joinedload(TimeEntry.task),
        )
        .order_by(TimeEntry.start_time.desc())
        .all()
    )

    try:
        # Get settings for delimiter
        settings = Settings.get_settings()
        delimiter = settings.export_delimiter

        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter)

        # Write header with task column
        writer.writerow(
            [
                "ID",
                "User",
                "Project",
                "Client",
                "Task",
                "Start Time",
                "End Time",
                "Duration (hours)",
                "Duration (formatted)",
                "Notes",
                "Tags",
                "Source",
                "Billable",
                "Created At",
                "Updated At",
            ]
        )

        # Write data (null-safe: user/project/client can be missing)
        for entry in entries:
            # Project.client is a property returning the client name string; use client_obj for the relationship
            client_name = (entry.client.name if entry.client else "") or (entry.project.client if entry.project else "")
            writer.writerow(
                [
                    entry.id,
                    (entry.user.display_name if entry.user else ""),
                    (entry.project.name if entry.project else ""),
                    client_name,
                    (entry.task.name if entry.task else ""),
                    entry.start_time.isoformat(),
                    entry.end_time.isoformat() if entry.end_time else "",
                    entry.duration_hours,
                    entry.duration_formatted,
                    entry.notes or "",
                    entry.tags or "",
                    entry.source,
                    "Yes" if entry.billable else "No",
                    entry.created_at.isoformat(),
                    entry.updated_at.isoformat() if entry.updated_at else "",
                ]
            )

        output.seek(0)

        # Create filename with filters indication
        filename_parts = [f"timetracker_export_{start_date}_to_{end_date}"]
        if project_id:
            filename_parts.append("project")
        if client_id:
            filename_parts.append("client")
        if task_id:
            filename_parts.append("task")
        filename = "_".join(filename_parts) + ".csv"

        # Track CSV export event with enhanced metadata
        log_event(
            "export.csv",
            user_id=current_user.id,
            export_type="time_entries",
            num_rows=len(entries),
            date_range_days=(end_dt - start_dt).days,
            filters_applied={
                "user_id": user_id,
                "project_id": project_id,
                "task_id": task_id,
                "client_id": client_id,
                "billable": billable,
                "source": source,
                "tags": tags,
            },
        )
        track_event(
            current_user.id,
            "export.csv",
            {
                "export_type": "time_entries",
                "num_rows": len(entries),
                "date_range_days": (end_dt - start_dt).days,
                "has_project_filter": project_id is not None,
                "has_client_filter": client_id is not None,
                "has_task_filter": task_id is not None,
                "has_billable_filter": billable is not None and billable != "all",
                "has_source_filter": source is not None and source != "all",
                "has_tags_filter": bool(tags),
            },
        )

        # Track performance
        try:
            duration_ms = (time.time() - start_time) * 1000
            csv_content = output.getvalue().encode("utf-8")
            track_export_performance(
                current_user.id,
                "csv",
                row_count=len(entries),
                duration_ms=duration_ms,
                file_size_bytes=len(csv_content),
            )
        except Exception:
            # Don't let tracking errors break the export
            pass

        record_report_generation_for_current_user()
        return send_file(io.BytesIO(csv_content), mimetype="text/csv", as_attachment=True, download_name=filename)
    except Exception:
        current_app.logger.exception("CSV export failed (reports.export_csv)")
        raise


@reports_bp.route("/reports/summary/export/pdf")
@login_required
@module_enabled("reports")
def export_summary_pdf():
    """Export summary report as a one-page PDF (today/week/month hours + top projects)."""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)
    today_hours = TimeEntry.get_total_hours_for_period(
        start_date=end_date.date(), user_id=current_user.id if not current_user.is_admin else None
    )
    week_hours = TimeEntry.get_total_hours_for_period(
        start_date=end_date.date() - timedelta(days=7), user_id=current_user.id if not current_user.is_admin else None
    )
    month_hours = TimeEntry.get_total_hours_for_period(
        start_date=start_date.date(), user_id=current_user.id if not current_user.is_admin else None
    )
    from app.utils.scope_filter import apply_project_scope_to_model

    scope_p = apply_project_scope_to_model(Project, current_user)
    projects_query = Project.query.filter_by(status="active")
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    elif not current_user.is_admin:
        time_entry_repo = TimeEntryRepository()
        project_ids = time_entry_repo.get_distinct_project_ids_for_user(current_user.id)
        projects_query = (
            projects_query.filter(Project.id.in_(project_ids))
            if project_ids
            else projects_query.filter(Project.id.in_([]))
        )
    projects = projects_query.all()
    project_stats = []
    for project in projects:
        hours = TimeEntry.get_total_hours_for_period(
            start_date=start_date.date(),
            project_id=project.id,
            user_id=current_user.id if not current_user.is_admin else None,
        )
        if hours > 0:
            project_stats.append({"project": project, "hours": hours})
    project_stats.sort(key=lambda x: x["hours"], reverse=True)
    project_stats = project_stats[:10]
    try:
        from app.utils.summary_report_pdf import build_summary_report_pdf

        pdf_bytes = build_summary_report_pdf(today_hours, week_hours, month_hours, project_stats)
    except Exception as e:
        current_app.logger.warning("Summary report PDF export failed: %s", e, exc_info=True)
        flash(_("PDF export failed: %(error)s", error=str(e)), "error")
        return redirect(url_for("reports.summary_report"))
    record_report_generation_for_current_user()
    filename = f"summary_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/summary")
@login_required
@module_enabled("reports")
def summary_report():
    """Summary report with key metrics"""
    # Get date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)

    # Get total hours for different periods
    today_hours = TimeEntry.get_total_hours_for_period(
        start_date=end_date.date(), user_id=current_user.id if not current_user.is_admin else None
    )

    week_hours = TimeEntry.get_total_hours_for_period(
        start_date=end_date.date() - timedelta(days=7), user_id=current_user.id if not current_user.is_admin else None
    )

    month_hours = TimeEntry.get_total_hours_for_period(
        start_date=start_date.date(), user_id=current_user.id if not current_user.is_admin else None
    )

    # Get top projects (scoped for subcontractors)
    from app.utils.scope_filter import apply_project_scope_to_model

    scope_p = apply_project_scope_to_model(Project, current_user)
    projects_query = Project.query.filter_by(status="active")
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    elif not current_user.is_admin:
        time_entry_repo = TimeEntryRepository()
        project_ids = time_entry_repo.get_distinct_project_ids_for_user(current_user.id)
        projects_query = (
            projects_query.filter(Project.id.in_(project_ids))
            if project_ids
            else projects_query.filter(Project.id.in_([]))
        )
    projects = projects_query.all()

    # Sort projects by total hours
    project_stats = []
    for project in projects:
        hours = TimeEntry.get_total_hours_for_period(
            start_date=start_date.date(),
            project_id=project.id,
            user_id=current_user.id if not current_user.is_admin else None,
        )
        if hours > 0:
            project_stats.append({"project": project, "hours": hours})

    project_stats.sort(key=lambda x: x["hours"], reverse=True)
    project_stats = project_stats[:10]  # Top 10 projects

    # Chart data: time by project (last 30 days)
    chart_labels_summary = [s["project"].name for s in project_stats]
    chart_hours_summary = [round(s["hours"], 2) for s in project_stats]

    # Daily trend for last 14 days (for line chart)
    from app.services import AnalyticsService

    analytics_service = AnalyticsService()
    trend_result = analytics_service.get_trends(
        user_id=current_user.id if not current_user.is_admin else None,
        days=14,
    )
    daily_trends_14d = trend_result.get("daily_trends", [])
    trend_dates = [t["date"] for t in daily_trends_14d]
    trend_hours = [t["hours"] for t in daily_trends_14d]

    return render_template(
        "reports/summary.html",
        today_hours=today_hours,
        week_hours=week_hours,
        month_hours=month_hours,
        project_stats=project_stats,
        chart_labels_summary=chart_labels_summary,
        chart_hours_summary=chart_hours_summary,
        trend_dates=trend_dates,
        trend_hours=trend_hours,
    )


@reports_bp.route("/reports/tasks")
@login_required
@module_enabled("reports")
def task_report():
    """Report of all tasks (completed and incomplete) with time entries logged within the date range, including hours spent per task"""
    project_id = request.args.get("project_id", type=int)
    user_id = request.args.get("user_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Filters data (scoped for subcontractors)
    from app.utils.scope_filter import apply_project_scope_to_model

    projects_query = Project.query.order_by(Project.name)
    scope_p = apply_project_scope_to_model(Project, current_user)
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    projects = projects_query.all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    # Default date range: last 30 days
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return render_template("reports/task_report.html", projects=projects, users=users)

    # Base tasks query: all tasks that have time entries within the date range
    tasks_query = Task.query.join(TimeEntry, TimeEntry.task_id == Task.id).filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    if project_id:
        tasks_query = tasks_query.filter(TimeEntry.project_id == project_id)

    # Optional: only tasks that have time entries by a specific user
    if user_id:
        tasks_query = tasks_query.filter(TimeEntry.user_id == user_id)

    # Get distinct task IDs (PostgreSQL requires ORDER BY cols in SELECT when using DISTINCT)
    task_ids_subq = tasks_query.with_entities(Task.id).distinct()
    tasks = (
        Task.query.options(joinedload(Task.project), joinedload(Task.assigned_user))
        .filter(Task.id.in_(task_ids_subq))
        .order_by(case((Task.status == "done", 0), else_=1), Task.name)
        .all()
    )
    task_ids = [t.id for t in tasks]

    # Single aggregation query for hours/entry count per task (avoids N+1)
    from app.repositories import TimeEntryRepository

    time_entry_repo = TimeEntryRepository()
    aggregates = time_entry_repo.get_task_aggregates(task_ids, start_dt, end_dt, project_id=project_id, user_id=user_id)
    agg_by_task = {tid: (total_sec, cnt) for tid, total_sec, cnt in aggregates}

    task_rows = []
    total_hours = 0.0
    for task in tasks:
        total_seconds, entries_count = agg_by_task.get(task.id, (0, 0))
        hours = round(total_seconds / 3600, 2)
        total_hours += hours
        task_rows.append(
            {
                "task": task,
                "project": task.project,
                "assignee": task.assigned_user,
                "status": task.status,
                "completed_at": task.completed_at,
                "hours": hours,
                "entries_count": entries_count,
            }
        )

    summary = {
        "tasks_count": len(task_rows),
        "total_hours": round(total_hours, 2),
    }

    return render_template(
        "reports/task_report.html",
        projects=projects,
        users=users,
        tasks=task_rows,
        summary=summary,
        start_date=start_date,
        end_date=end_date,
        selected_project=project_id,
        selected_user=user_id,
    )


def _time_entries_report_query(request, require_dates=True, return_query=False):
    """Shared query logic for time entries report and its exports.
    When return_query=False: returns (entries, start_dt, end_dt, start_date, end_date) or (None, None, None, start_date, end_date) on date error.
    When return_query=True: returns (query, start_dt, end_dt, start_date, end_date) with query having filters applied (no options/order_by/execution).
    """
    from app.utils.client_lock import enforce_locked_client_id

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)
    client_id = request.args.get("client_id", type=int)
    client_id = enforce_locked_client_id(client_id)
    task_id = request.args.get("task_id", type=int)
    billed = request.args.get("billed", "all")  # 'all', 'yes', 'no'

    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        if require_dates:
            return None, None, None, start_date, end_date
        start_dt = datetime.utcnow() - timedelta(days=30)
        end_dt = datetime.utcnow()

    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None),
        TimeEntry.start_time >= start_dt,
        TimeEntry.start_time <= end_dt,
    )

    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if user_id:
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            return None, None, None, start_date, end_date

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)
    if task_id:
        query = query.filter(TimeEntry.task_id == task_id)
    if billed == "yes":
        query = query.filter(TimeEntry.paid == True)
    elif billed == "no":
        query = query.filter(TimeEntry.paid == False)

    if client_id:
        project_ids_for_client = db.session.query(Project.id).filter(Project.client_id == client_id)
        query = query.filter(or_(TimeEntry.client_id == client_id, TimeEntry.project_id.in_(project_ids_for_client)))

    # Subcontractor scope: restrict to allowed projects
    from app.utils.scope_filter import get_allowed_project_ids

    allowed_project_ids = get_allowed_project_ids(current_user)
    if allowed_project_ids is not None:
        if not allowed_project_ids:
            query = query.filter(TimeEntry.project_id.in_([]))
        else:
            query = query.filter(TimeEntry.project_id.in_(allowed_project_ids))

    if return_query:
        return query, start_dt, end_dt, start_date, end_date

    entries = (
        query.options(
            joinedload(TimeEntry.project).joinedload(Project.client_obj),
            joinedload(TimeEntry.user),
            joinedload(TimeEntry.task),
            joinedload(TimeEntry.client),
        )
        .order_by(TimeEntry.start_time.desc())
        .all()
    )

    return entries, start_dt, end_dt, start_date, end_date


@reports_bp.route("/reports/time-entries")
@login_required
@module_enabled("reports")
def time_entries_report():
    """Time Entries report: list all time entries (billed and unbilled) with Date, Start, Stop, Duration, Project, Task, Notes, Billed, Client."""
    from app.utils.client_lock import enforce_locked_client_id

    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    user_id_arg = request.args.get("user_id", type=int)
    if not can_view_all and user_id_arg and user_id_arg != current_user.id:
        flash(_("You do not have permission to view other users' time entries"), "error")
        return redirect(url_for("reports.time_entries_report"))

    from app.utils.scope_filter import apply_client_scope_to_model, apply_project_scope_to_model

    projects_query = Project.query.filter_by(status="active").order_by(Project.name)
    scope_p = apply_project_scope_to_model(Project, current_user)
    if scope_p is not None:
        projects_query = projects_query.filter(scope_p)
    projects = projects_query.all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    clients_query = Client.query.filter_by(status="active").order_by(Client.name)
    scope_c = apply_client_scope_to_model(Client, current_user)
    if scope_c is not None:
        clients_query = clients_query.filter(scope_c)
    clients = clients_query.all()
    tasks = Task.query.order_by(Task.name).all()

    base_query, start_dt, end_dt, start_date, end_date = _time_entries_report_query(
        request, require_dates=True, return_query=True
    )
    if base_query is None:
        flash(_("Invalid date format"), "error")
        return render_template(
            "reports/time_entries_report.html",
            projects=projects,
            users=users,
            clients=clients,
            tasks=tasks,
            entries=[],
            summary={"entries_count": 0, "total_hours": 0},
            start_date=start_date,
            end_date=end_date,
            selected_user=request.args.get("user_id", type=int),
            selected_project=request.args.get("project_id", type=int),
            selected_client=enforce_locked_client_id(request.args.get("client_id", type=int)),
            selected_task=request.args.get("task_id", type=int),
            selected_billed=request.args.get("billed", "all"),
            pagination=None,
        )

    from app.constants import DEFAULT_PAGE_SIZE, MAX_PAGE_SIZE

    page = request.args.get("page", 1, type=int)
    per_page = min(request.args.get("per_page", DEFAULT_PAGE_SIZE, type=int), MAX_PAGE_SIZE)
    per_page = max(1, per_page)

    # Summary from aggregation (same filters, not just current page)
    total_count = base_query.with_entities(func.count(TimeEntry.id)).scalar() or 0
    total_seconds = base_query.with_entities(func.sum(TimeEntry.duration_seconds)).scalar() or 0
    summary = {"entries_count": total_count, "total_hours": round((total_seconds or 0) / 3600, 2)}

    entries_query = base_query.options(
        joinedload(TimeEntry.project).joinedload(Project.client_obj),
        joinedload(TimeEntry.user),
        joinedload(TimeEntry.task),
        joinedload(TimeEntry.client),
    ).order_by(TimeEntry.start_time.desc())
    paginated = entries_query.paginate(page=page, per_page=per_page, error_out=False)
    entries = paginated.items

    pagination = {
        "page": paginated.page,
        "per_page": paginated.per_page,
        "total": paginated.total,
        "pages": paginated.pages,
        "has_next": paginated.has_next,
        "has_prev": paginated.has_prev,
        "next_page": paginated.page + 1 if paginated.has_next else None,
        "prev_page": paginated.page - 1 if paginated.has_prev else None,
    }

    return render_template(
        "reports/time_entries_report.html",
        projects=projects,
        users=users,
        clients=clients,
        tasks=tasks,
        entries=entries,
        summary=summary,
        start_date=start_date,
        end_date=end_date,
        selected_user=request.args.get("user_id", type=int),
        selected_project=request.args.get("project_id", type=int),
        selected_client=enforce_locked_client_id(request.args.get("client_id", type=int)),
        selected_task=request.args.get("task_id", type=int),
        selected_billed=request.args.get("billed", "all"),
        pagination=pagination,
    )


@reports_bp.route("/reports/time-entries/export/excel")
@login_required
@module_enabled("reports")
def time_entries_export_excel():
    """Export Time Entries report as Excel (same filters as report, includes Billed and Client)."""
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    user_id_arg = request.args.get("user_id", type=int)
    if not can_view_all and user_id_arg and user_id_arg != current_user.id:
        flash(_("You do not have permission to export other users' time entries"), "error")
        return redirect(url_for("reports.time_entries_report"))

    entries, start_dt, end_dt, start_date, end_date = _time_entries_report_query(request, require_dates=True)
    if entries is None:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.time_entries_report"))

    columns = ["date", "start_time", "end_time", "duration_hours", "project", "task", "notes", "billed", "client"]
    if can_view_all:
        columns.insert(2, "user")  # insert user after end_time for multi-user export
    output, filename = create_time_entries_excel(entries, filename_prefix="time_entries_report", columns=columns)
    log_event(
        "export.excel",
        user_id=current_user.id,
        export_type="time_entries_report",
        num_rows=len(entries),
        date_range_days=(end_dt - start_dt).days,
    )
    track_event(
        current_user.id,
        "export.excel",
        {"export_type": "time_entries_report", "num_rows": len(entries)},
    )
    record_report_generation_for_current_user()
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/time-entries/export/csv")
@login_required
@module_enabled("reports")
def time_entries_export_csv():
    """Export Time Entries report as CSV (same filters as report, includes Billed and Client)."""
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    user_id_arg = request.args.get("user_id", type=int)
    if not can_view_all and user_id_arg and user_id_arg != current_user.id:
        flash(_("You do not have permission to export other users' time entries"), "error")
        return redirect(url_for("reports.time_entries_report"))

    entries, start_dt, end_dt, start_date, end_date = _time_entries_report_query(request, require_dates=True)
    if entries is None:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.time_entries_report"))

    settings = Settings.get_settings()
    delimiter = settings.export_delimiter
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter)
    headers = [
        _("Date"),
        _("Start"),
        _("End"),
        _("Duration (hours)"),
        _("Project"),
        _("Task"),
        _("Notes"),
        _("Billed"),
        _("Client"),
    ]
    if can_view_all:
        headers.insert(2, _("User"))  # after End
    writer.writerow(headers)
    for entry in entries:
        client_name = (
            (entry.client.name if entry.client else "") or (entry.project.client if entry.project else "") or ""
        )
        row = [
            entry.start_time.date().isoformat() if entry.start_time else "",
            entry.start_time.isoformat() if entry.start_time else "",
            entry.end_time.isoformat() if entry.end_time else "",
            entry.duration_hours if entry.end_time else "",
            entry.project.name if entry.project else "",
            entry.task.name if entry.task else "",
            entry.notes or "",
            _("Yes") if entry.paid else _("No"),
            client_name,
        ]
        if can_view_all:
            row.insert(2, entry.user.display_name if entry.user else "Unknown")
        writer.writerow(row)
    output.seek(0)
    filename = f"time_entries_report_{start_date}_to_{end_date}.csv"
    record_report_generation_for_current_user()
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/export/excel")
@login_required
@module_enabled("reports")
def export_excel():
    """Export time entries as Excel file"""
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.reports"))

    # Get time entries
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if user_id:
        # Only allow filtering by other users if they have permission
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            flash(_("You do not have permission to export other users' time entries"), "error")
            return redirect(url_for("reports.reports"))

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)

    entries = (
        query.options(
            joinedload(TimeEntry.project).joinedload(Project.client_obj),
            joinedload(TimeEntry.user),
            joinedload(TimeEntry.task),
        )
        .order_by(TimeEntry.start_time.desc())
        .all()
    )

    # Create Excel file
    output, filename = create_time_entries_excel(entries, filename_prefix="timetracker_export")

    # Track Excel export event
    log_event(
        "export.excel",
        user_id=current_user.id,
        export_type="time_entries",
        num_rows=len(entries),
        date_range_days=(end_dt - start_dt).days,
    )
    track_event(
        current_user.id,
        "export.excel",
        {"export_type": "time_entries", "num_rows": len(entries), "date_range_days": (end_dt - start_dt).days},
    )

    record_report_generation_for_current_user()
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/project/export/excel")
@login_required
@module_enabled("reports")
def export_project_excel():
    """Export project report as Excel file"""
    project_id = request.args.get("project_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    user_id = request.args.get("user_id", type=int)

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.project_report"))

    # Get time entries
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)

    if user_id:
        # Only allow filtering by other users if they have permission
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            flash(_("You do not have permission to export other users' time entries"), "error")
            return redirect(url_for("reports.project_report"))

    entries = query.all()

    # Aggregate by project
    projects_map = {}
    for entry in entries:
        project = entry.project
        if not project:
            continue
        if project.id not in projects_map:
            projects_map[project.id] = {
                "name": project.name,
                "client": project.client if project.client else "",
                "total_hours": 0,
                "billable_hours": 0,
                "hourly_rate": float(project.hourly_rate) if project.hourly_rate else 0,
                "billable_amount": 0,
                "total_costs": 0,
                "total_value": 0,
            }
        agg = projects_map[project.id]
        hours = entry.duration_hours
        agg["total_hours"] += hours
        if entry.billable and project.billable:
            agg["billable_hours"] += hours
            if project.hourly_rate:
                agg["billable_amount"] += hours * float(project.hourly_rate)

    projects_data = list(projects_map.values())

    from app.telemetry.otel_setup import business_span, record_export_duration_seconds

    _exp_t0 = time.monotonic()
    with business_span(
        "report.export",
        user_id=current_user.id,
        export_type="project_report",
        num_projects=len(projects_data),
    ):
        output, filename = create_project_report_excel(projects_data, start_date, end_date)
    record_export_duration_seconds(time.monotonic() - _exp_t0, "project_report")

    # Track event
    log_event("export.excel", user_id=current_user.id, export_type="project_report", num_projects=len(projects_data))
    track_event(current_user.id, "export.excel", {"export_type": "project_report", "num_projects": len(projects_data)})

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/user/export/excel")
@login_required
@module_enabled("reports")
def export_user_excel():
    """Export user report as Excel file"""
    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.user_report"))

    # Get time entries
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if user_id:
        # Only allow filtering by other users if they have permission
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            flash(_("You do not have permission to export other users' time entries"), "error")
            return redirect(url_for("reports.reports"))

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)

    entries = (
        query.options(
            joinedload(TimeEntry.user),
        )
        .order_by(TimeEntry.start_time.desc())
        .all()
    )

    # Group by user
    user_totals = {}
    for entry in entries:
        username = entry.user.display_name if entry.user else "Unknown"
        if username not in user_totals:
            user_totals[username] = {
                "hours": 0,
                "billable_hours": 0,
                "user_obj": entry.user,
            }
        user_totals[username]["hours"] += entry.duration_hours
        if entry.billable:
            user_totals[username]["billable_hours"] += entry.duration_hours

    # Calculate overtime
    from app.utils.overtime import calculate_period_overtime

    for username, data in user_totals.items():
        if data["user_obj"]:
            overtime_data = calculate_period_overtime(data["user_obj"], start_dt.date(), end_dt.date())
            data["regular_hours"] = overtime_data["regular_hours"]
            data["overtime_hours"] = overtime_data["overtime_hours"]
            data["undertime_hours"] = overtime_data.get("undertime_hours", 0)
            data["days_under"] = overtime_data.get("days_under", 0)
            data["days_with_overtime"] = overtime_data["days_with_overtime"]
        else:
            data["regular_hours"] = data["hours"]
            data["overtime_hours"] = 0
            data["undertime_hours"] = 0
            data["days_under"] = 0
            data["days_with_overtime"] = 0

    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "User Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"User Report: {start_date} to {end_date}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "User",
        "Total Hours",
        "Regular Hours",
        "Overtime Hours",
        "Undertime Hours",
        "Billable Hours",
        "Days with Overtime",
        "Days Under",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    row_num = 4
    for username, data in sorted(user_totals.items()):
        ws.cell(row=row_num, column=1).value = username
        ws.cell(row=row_num, column=2).value = round(data["hours"], 2)
        ws.cell(row=row_num, column=3).value = round(data.get("regular_hours", data["hours"]), 2)
        ws.cell(row=row_num, column=4).value = round(data.get("overtime_hours", 0), 2)
        ws.cell(row=row_num, column=5).value = round(data.get("undertime_hours", 0), 2)
        ws.cell(row=row_num, column=6).value = round(data["billable_hours"], 2)
        ws.cell(row=row_num, column=7).value = data.get("days_with_overtime", 0)
        ws.cell(row=row_num, column=8).value = data.get("days_under", 0)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num > 1:
                cell.number_format = "0.00"

        row_num += 1

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = max(len(header), 15)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"user_report_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    log_event("export.excel", user_id=current_user.id, export_type="user_report", num_users=len(user_totals))
    track_event(current_user.id, "export.excel", {"export_type": "user_report", "num_users": len(user_totals)})

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/user/export/entries/excel")
@login_required
@module_enabled("reports")
def export_user_entries_excel():
    """Export detailed user report as Excel (one row per time entry)."""
    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Columns are customizable via repeated query params: ?columns=date&columns=user...
    columns = [c.strip() for c in request.args.getlist("columns") if (c or "").strip()]
    if not columns:
        # Default matches issue #483 request (Excel-friendly)
        columns = ["date", "user", "project", "task", "duration_hours", "notes"]

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.user_report"))

    # Get time entries
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    query = TimeEntry.query.filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    if user_id:
        # Only allow filtering by other users if they have permission
        if can_view_all:
            query = query.filter(TimeEntry.user_id == user_id)
        elif user_id != current_user.id:
            flash(_("You do not have permission to export other users' time entries"), "error")
            return redirect(url_for("reports.reports"))

    if project_id:
        query = query.filter(TimeEntry.project_id == project_id)

    entries = (
        query.options(
            joinedload(TimeEntry.project).joinedload(Project.client_obj),
            joinedload(TimeEntry.user),
            joinedload(TimeEntry.task),
        )
        .order_by(TimeEntry.start_time.desc())
        .all()
    )

    # Create Excel file (row-per-entry)
    output, filename = create_time_entries_excel(entries, filename_prefix="user_entries", columns=columns)

    log_event(
        "export.excel",
        user_id=current_user.id,
        export_type="user_entries",
        num_rows=len(entries),
        filters_applied={"user_id": user_id, "project_id": project_id, "start_date": start_date, "end_date": end_date},
        columns=columns,
    )
    track_event(
        current_user.id,
        "export.excel",
        {"export_type": "user_entries", "num_rows": len(entries), "columns": columns},
    )

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/task/export/excel")
@login_required
@module_enabled("reports")
def export_task_excel():
    """Export task report as Excel file - includes all tasks (completed and incomplete) with time entries in date range"""
    project_id = request.args.get("project_id", type=int)
    user_id = request.args.get("user_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.task_report"))

    # Get tasks: all tasks that have time entries within the date range (eager load to avoid N+1)
    tasks_query = Task.query.join(TimeEntry, TimeEntry.task_id == Task.id).filter(
        TimeEntry.end_time.isnot(None), TimeEntry.start_time >= start_dt, TimeEntry.start_time <= end_dt
    )

    if project_id:
        tasks_query = tasks_query.filter(TimeEntry.project_id == project_id)

    if user_id:
        tasks_query = tasks_query.filter(TimeEntry.user_id == user_id)

    task_ids_subq = tasks_query.with_entities(Task.id).distinct()
    tasks = (
        Task.query.options(joinedload(Task.project), joinedload(Task.assigned_user))
        .filter(Task.id.in_(task_ids_subq))
        .order_by(case((Task.status == "done", 0), else_=1), Task.name)
        .all()
    )
    task_ids = [t.id for t in tasks]

    from app.repositories import TimeEntryRepository

    time_entry_repo = TimeEntryRepository()
    aggregates = time_entry_repo.get_task_aggregates(task_ids, start_dt, end_dt, project_id=project_id, user_id=user_id)
    agg_by_task = {tid: (int(total_sec or 0), cnt) for tid, total_sec, cnt in aggregates}

    task_rows = []
    for task in tasks:
        total_seconds, _entry_count = agg_by_task.get(task.id, (0, 0))
        hours = round(total_seconds / 3600, 2)
        task_rows.append(
            {
                "task": task,
                "project": task.project,
                "status": task.status,
                "completed_at": task.completed_at,
                "hours": hours,
            }
        )

    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Task Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Task Report: {start_date} to {end_date}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Task", "Project", "Status", "Completed At", "Hours"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    row_num = 4
    for row_data in task_rows:
        ws.cell(row=row_num, column=1).value = row_data["task"].name
        ws.cell(row=row_num, column=2).value = row_data["project"].name if row_data["project"] else "N/A"
        ws.cell(row=row_num, column=3).value = row_data["status"].replace("_", " ").title()
        ws.cell(row=row_num, column=4).value = (
            row_data["completed_at"].strftime("%Y-%m-%d") if row_data["completed_at"] else "N/A"
        )
        ws.cell(row=row_num, column=5).value = row_data["hours"]

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num == 5:
                cell.number_format = "0.00"

        row_num += 1

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = max(len(header), 15)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"task_report_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    log_event("export.excel", user_id=current_user.id, export_type="task_report", num_tasks=len(task_rows))
    track_event(current_user.id, "export.excel", {"export_type": "task_report", "num_tasks": len(task_rows)})

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/unpaid-hours")
@login_required
@module_enabled("reports")
def unpaid_hours_report():
    """Report showing unpaid hours per client"""
    from app.utils.client_lock import enforce_locked_client_id

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    client_id = request.args.get("client_id", type=int)
    client_id = enforce_locked_client_id(client_id)

    # Get clients for filter (scoped for subcontractors)
    from app.utils.scope_filter import apply_client_scope_to_model

    clients_query = Client.query.filter_by(status="active").order_by(Client.name)
    scope_c = apply_client_scope_to_model(Client, current_user)
    if scope_c is not None:
        clients_query = clients_query.filter(scope_c)
    clients = clients_query.all()
    only_one_client = len(clients) == 1
    single_client = clients[0] if only_one_client else None

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return render_template(
            "reports/unpaid_hours_report.html",
            clients=clients,
            only_one_client=only_one_client,
            single_client=single_client,
        )

    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    from app.services import ReportingService

    data = ReportingService().get_unpaid_hours_report_data(
        start_dt=start_dt,
        end_dt=end_dt,
        client_id=client_id,
        current_user_id=current_user.id,
        can_view_all=can_view_all,
    )
    client_data = data["client_data"]
    summary = data["summary"]

    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.args.get("format") == "json":
        return jsonify(
            {
                "summary": summary,
                "client_data": [
                    {
                        "client_id": d["client"].id,
                        "client_name": d["client"].name,
                        "client_email": getattr(d["client"], "email", None),
                        "total_hours": d["total_hours"],
                        "billable_hours": d["billable_hours"],
                        "estimated_amount": d["estimated_amount"],
                        "projects": [
                            {
                                "project_id": p["project"].id,
                                "project_name": p["project"].name,
                                "hours": p["hours"],
                                "rate": p["rate"],
                            }
                            for p in d["projects"]
                        ],
                        "entries": [
                            {
                                "id": e.id,
                                "user": e.user.display_name if e.user else "Unknown",
                                "project": e.project.name if e.project else "No Project",
                                "task": e.task.name if e.task else None,
                                "start_time": e.start_time.isoformat() if e.start_time else None,
                                "end_time": e.end_time.isoformat() if e.end_time else None,
                                "duration_hours": round(e.duration_hours, 2),
                                "notes": e.notes or "",
                            }
                            for e in d["entries"]
                        ],
                    }
                    for d in client_data
                ],
            }
        )

    return render_template(
        "reports/unpaid_hours_report.html",
        clients=clients,
        only_one_client=only_one_client,
        single_client=single_client,
        client_data=client_data,
        summary=summary,
        start_date=start_date,
        end_date=end_date,
        selected_client=client_id,
    )


@reports_bp.route("/reports/unpaid-hours/export/excel")
@login_required
@module_enabled("reports")
def export_unpaid_hours_excel():
    """Export unpaid hours report as Excel file, organized by project"""
    from app.utils.client_lock import enforce_locked_client_id

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    client_id = request.args.get("client_id", type=int)
    client_id = enforce_locked_client_id(client_id)

    # Parse dates
    if not start_date:
        start_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.utcnow().strftime("%Y-%m-%d")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash(_("Invalid date format"), "error")
        return redirect(url_for("reports.unpaid_hours_report"))

    # Get all billable time entries in the date range
    can_view_all = current_user.is_admin or current_user.has_permission("view_all_time_entries")
    from sqlalchemy.orm import joinedload

    query = TimeEntry.query.options(
        joinedload(TimeEntry.user),
        joinedload(TimeEntry.project),
        joinedload(TimeEntry.task),
        joinedload(TimeEntry.client),
    ).filter(
        TimeEntry.end_time.isnot(None),
        TimeEntry.billable == True,
        TimeEntry.start_time >= start_dt,
        TimeEntry.start_time <= end_dt,
    )

    # Filter by user if no permission to view all
    if not can_view_all:
        query = query.filter(TimeEntry.user_id == current_user.id)

    all_entries = query.all()

    # Filter by client if specified (check both entry.client_id and project.client_id)
    if client_id:
        all_entries = [
            e for e in all_entries if (e.client_id == client_id) or (e.project and e.project.client_id == client_id)
        ]

    # Get all invoice items to check which time entries are already invoiced
    from app.models.invoice import InvoiceItem

    all_invoice_items = (
        InvoiceItem.query.join(Invoice)
        .filter(InvoiceItem.time_entry_ids.isnot(None), InvoiceItem.time_entry_ids != "")
        .all()
    )

    # Build a set of time entry IDs that are in fully paid invoices
    billed_entry_ids = set()

    for item in all_invoice_items:
        if not item.time_entry_ids:
            continue
        entry_ids = [int(eid.strip()) for eid in item.time_entry_ids.split(",") if eid.strip().isdigit()]
        invoice = item.invoice
        if invoice and invoice.payment_status == "fully_paid":
            billed_entry_ids.update(entry_ids)

    # Filter entries: only include those that are NOT in fully paid invoices
    unpaid_entries = [e for e in all_entries if e.id not in billed_entry_ids]

    # Debug: Check if we have any entries
    if not unpaid_entries:
        # Still create a file with empty data to show the issue
        pass

    # Group by project
    project_data = {}
    for entry in unpaid_entries:
        # Get project
        project = entry.project
        if not project:
            continue

        project_id = project.id
        if project_id not in project_data:
            # Get client from entry or from project
            client = None
            if entry.client_id:
                client = entry.client
            elif project.client_id:
                client = project.client_obj

            project_data[project_id] = {
                "project": project,
                "client": client,
                "entries": [],
                "total_hours": 0.0,
                "estimated_amount": 0.0,
            }

        hours = entry.duration_hours
        project_data[project_id]["total_hours"] += hours
        project_data[project_id]["entries"].append(entry)

        # Calculate estimated amount
        rate = 0.0
        if project.hourly_rate:
            rate = float(project.hourly_rate)
        elif client and client.default_hourly_rate:
            rate = float(client.default_hourly_rate)
        project_data[project_id]["estimated_amount"] += hours * rate

    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.merge_cells("A1:D1")
    title_cell = summary_ws["A1"]
    title_cell.value = f"Unpaid Hours Report: {start_date} to {end_date}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    # Summary headers
    summary_headers = ["Client", "Project", "Total Hours", "Estimated Amount"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Summary data
    row_num = 4
    total_hours = 0.0
    total_amount = 0.0

    if not project_data:
        # No data message
        summary_ws.cell(row=4, column=1, value="No unpaid hours found for the selected period.")
        summary_ws.merge_cells("A4:D4")
    else:
        for project_id, data in sorted(
            project_data.items(),
            key=lambda x: (x[1]["client"].name if x[1]["client"] and x[1]["client"].name else "", x[1]["project"].name),
        ):
            summary_ws.cell(row=row_num, column=1, value=data["client"].name if data["client"] else "N/A").border = (
                border
            )
            summary_ws.cell(row=row_num, column=2, value=data["project"].name).border = border
            summary_ws.cell(row=row_num, column=3, value=round(data["total_hours"], 2)).border = border
            summary_ws.cell(row=row_num, column=3).number_format = "0.00"
            summary_ws.cell(row=row_num, column=4, value=round(data["estimated_amount"], 2)).border = border
            summary_ws.cell(row=row_num, column=4).number_format = "0.00"
            total_hours += data["total_hours"]
            total_amount += data["estimated_amount"]
            row_num += 1

        # Summary totals
        row_num += 1
        summary_ws.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
        summary_ws.cell(row=row_num, column=2, value="").font = Font(bold=True)
        summary_ws.cell(row=row_num, column=3, value=round(total_hours, 2)).font = Font(bold=True)
        summary_ws.cell(row=row_num, column=3).number_format = "0.00"
        summary_ws.cell(row=row_num, column=4, value=round(total_amount, 2)).font = Font(bold=True)
        summary_ws.cell(row=row_num, column=4).number_format = "0.00"

    # Auto-adjust column widths for summary
    for col_idx in range(1, len(summary_headers) + 1):
        column = get_column_letter(col_idx)
        summary_ws.column_dimensions[column].width = 20

    # Create a sheet for each project
    if project_data:
        for project_id, data in sorted(
            project_data.items(),
            key=lambda x: (x[1]["client"].name if x[1]["client"] and x[1]["client"].name else "", x[1]["project"].name),
        ):
            project = data["project"]
            client = data["client"]

            # Create sheet name (Excel has 31 char limit for sheet names)
            sheet_name = f"{client.name[:15]}-{project.name[:15]}" if client else project.name[:31]
            sheet_name = (
                sheet_name.replace("/", "-")
                .replace("\\", "-")
                .replace("?", "-")
                .replace("*", "-")
                .replace("[", "-")
                .replace("]", "-")
                .replace(":", "-")
            )

            ws = wb.create_sheet(sheet_name)

            # Title
            ws.merge_cells("A1:G1")
            title_cell = ws["A1"]
            title_cell.value = f"{client.name if client else 'N/A'} - {project.name}"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center")

            # Project info
            ws.cell(row=2, column=1, value="Client:").font = Font(bold=True)
            ws.cell(row=2, column=2, value=client.name if client else "N/A")
            ws.cell(row=3, column=1, value="Project:").font = Font(bold=True)
            ws.cell(row=3, column=2, value=project.name)
            ws.cell(row=4, column=1, value="Total Hours:").font = Font(bold=True)
            ws.cell(row=4, column=2, value=round(data["total_hours"], 2))
            ws.cell(row=4, column=2).number_format = "0.00"
            ws.cell(row=5, column=1, value="Estimated Amount:").font = Font(bold=True)
            ws.cell(row=5, column=2, value=round(data["estimated_amount"], 2))
            ws.cell(row=5, column=2).number_format = "0.00"

            # Headers
            headers = ["User", "Task", "Date", "Start Time", "End Time", "Duration (hours)", "Notes"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Data rows
            row_num = 8
            for entry in sorted(data["entries"], key=lambda x: x.start_time if x.start_time else datetime.min):
                ws.cell(row=row_num, column=1, value=entry.user.display_name if entry.user else "Unknown").border = (
                    border
                )
                ws.cell(row=row_num, column=2, value=entry.task.name if entry.task else "-").border = border
                ws.cell(
                    row=row_num, column=3, value=entry.start_time.strftime("%Y-%m-%d") if entry.start_time else "-"
                ).border = border
                ws.cell(
                    row=row_num, column=4, value=entry.start_time.strftime("%H:%M:%S") if entry.start_time else "-"
                ).border = border
                ws.cell(
                    row=row_num, column=5, value=entry.end_time.strftime("%H:%M:%S") if entry.end_time else "-"
                ).border = border
                ws.cell(row=row_num, column=6, value=round(entry.duration_hours, 2)).border = border
                ws.cell(row=row_num, column=6).number_format = "0.00"
                ws.cell(row=row_num, column=7, value=entry.notes or "-").border = border
                row_num += 1

            # Totals row
            row_num += 1
            ws.cell(row=row_num, column=5, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=row_num, column=6, value=round(data["total_hours"], 2)).font = Font(bold=True)
            ws.cell(row=row_num, column=6).number_format = "0.00"

            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                column = get_column_letter(col_idx)
                max_length = 15
                for row in ws.iter_rows(min_row=7, max_row=row_num, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError) as e:
                            # Cell value may be None or not have expected attributes
                            current_app.logger.debug(f"Error reading cell value: {e}")
                            pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"unpaid_hours_report_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Track event
    log_event(
        "export.excel", user_id=current_user.id, export_type="unpaid_hours_report", num_projects=len(project_data)
    )
    track_event(
        current_user.id, "export.excel", {"export_type": "unpaid_hours_report", "num_projects": len(project_data)}
    )

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
